#print all columns name openpyxl

import openpyxl

path = " path of file"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
max_col = sheet_obj.max_column

for i in range(1, max_col +1):
    cell_obj = sheet_obj.cell(row = 1, column = i)
    print(cell_obj.value)