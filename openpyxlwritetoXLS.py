#program to write to an excel with openpyxl

import openpyxl

wb = openpxl.Workbook()

sheet = wb.active

c1 = sheet.cell(row = 1, column = 1)
c1.value = "Michael"

c2.sheet.cell(row = 1, column = 2)
c2.value = "Sammons"

wb.save("file path to be used")